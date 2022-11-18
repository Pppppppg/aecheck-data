from openpyxl import load_workbook
from function import add_data_to_sheet
from config import FILE_NAME

workbook = load_workbook(FILE_NAME, data_only=True)


"""
캐릭터 이름 번역

퍼스널리티 시트의 1행 (A행) 참조
"""
per_sheet = workbook["퍼스널리티"]
char_list = [i[0].value for i in per_sheet.iter_rows(min_row=2)]
add_data_to_sheet("캐릭터", workbook["캐릭번역"], char_list)


"""
버디 이름 번역

버디 시트의 2행 (B행) 참조
"""
per_sheet = workbook["버디"]
buddy_list = [i[1].value for i in per_sheet.iter_rows(min_row=2)]
add_data_to_sheet("버디", workbook["캐릭번역"], buddy_list)


"""
직업서 번역

캐릭터 시트의 13행 (M행) 참조
"""
char_sheet = workbook["캐릭터"]
book_list = [i[12].value for i in char_sheet.iter_rows(min_row=2)]
add_data_to_sheet("직업서", workbook["캐릭번역"], book_list)


"""
특성 번역

퍼스널리티 시트의 3행 (C행) 참조
"""
per_list = []
per_sheet = workbook["퍼스널리티"]
for row in per_sheet.iter_rows(min_row=2):
    per_list += row[2].value.split(",")

add_data_to_sheet("퍼스널리티", workbook["특성번역"], per_list)



workbook.save(FILE_NAME)
print("Save Excel. COMPLETE")